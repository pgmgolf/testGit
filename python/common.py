# -*- coding: utf-8 -*-

from flask import Flask, session, redirect, url_for
from flask import Flask, request, jsonify, current_app, abort, send_from_directory, send_file
from flask_cors import CORS, cross_origin
from flask import Blueprint
from flaskext.mysql import MySQL
from pymongo import MongoClient
import json
from bson import BSON
from bson import json_util
import os
import pymssql
import datetime
import re
import random
import string
import math
import base64
import jwt
import codecs
import hashlib
from flask import Flask, request, make_response
from pyreportjasper import JasperPy
from openpyxl import load_workbook
from werkzeug import secure_filename

from db_config import *
common = Blueprint('common', __name__)

def GetPathFlask():
    return os.path.dirname(os.path.abspath(__file__))

def Encrypt(str):
    strEncry = ""
    if len(str) == 0:
        return ""
    for i in range(0, len(str), 1):
        n = ord(str[i]) + i
        n1 = math.floor(n / 16) + 48
        if (n % 16) > 9:
            n2 = (n % 16) + 55
        else:
            n2 = (n % 16) + 48
        strEncry = strEncry + chr(int(n1)) + chr(int(n2))
    return strEncry

def Decrypt(str):
    strEncry = ""
    if len(str) == 0:
        return ""
    for i in range(0, len(str), 2):
        n = (ord(str[i]) - 48) * 16
        if i < len(str) - 1:
            if ord(str[i + 1]) > 57:
                n = n + ((ord(str[i + 1]) - 55) - (i / 2))
            else:
                n = n + ((ord(str[i + 1]) - 48) - (i / 2))

        strEncry = strEncry + chr(int(n))
    return strEncry

def pymssqlCustomExceptionMessage(e):
    return str(e.args[1])[2:str(e.args[1]).find('DB-Lib error message')]

def randomStringDigits(stringLength=20):
    """Generate a random string of letters and digits """
    lettersAndDigits = string.ascii_letters + string.digits
    return ''.join(random.choice(lettersAndDigits) for i in range(stringLength))

def encryptHashSHA512(hash_string):
    sha_signature = \
        hashlib.sha512(hash_string.encode()).hexdigest()
    return sha_signature
    sha_signature = encrypt_string(hash_string)
    return sha_signature

def valid_email(email):
  return bool(re.search(r"^[\w\.\+\-]+\@[\w]+\.[a-z]{2,3}$", email))

def toJson(data,columns):
    results = []
    for row in data:
        results.append(dict(zip(columns, row)))
    return results

def toJsonOne(data,columns):
    results = {}
    for row in data:
        results = dict(zip(columns, row))
    return results

def num_to_col_letters(num):
    letters = ''
    while num:
        mod = (num - 1) % 26
        letters += chr(mod + 65)
        num = (num - 1) // 26
    return ''.join(reversed(letters))

def existsListString(listString,value):
    for str in listString:
        if value == str:
            return True

    return False

def ifNoneValue(value,DefaultValue):
    return value if value != None else DefaultValue

# def getUserToken(token):
#     try:
#         if token is not None:
#             token = token.replace("Bearer-", "")
#         return jwt.decode(str(token), current_app.config['SECRET_KEY'], 'HS256')
#     except Exception as e:
#         raise e

def getUserToken(token):
    try:
        isSuccess = True

        if token is None:
            raise ValueError("Authorization Token is none.")

        if len(token) == 0:
            raise ValueError("Authorization Token is empty.")

        if token is not None:
            token = token.replace("Bearer-", "")

        payload = jwt.decode(str(token), current_app.config['SECRET_KEY'], 'HS256')

        expireUserDate = datetime.datetime.fromtimestamp(payload["exp"])

        if expireUserDate < datetime.datetime.now():
            raise ValueError("Authorization Token Expire " + expireUserDate.strftime("%Y-%m-%d %H:%M:%S.%f"))

        columns = ['user_id','iat','exp']
        data = [(payload['user_id'],payload['iat'],payload['exp'])]

        displayColumns = ['isSuccess','data']
        displayData = [(isSuccess,toJsonOne(data,columns))]

        return toJsonOne(displayData,displayColumns)
    except Exception as e:
        isSuccess = False
        reasonCode = 500
        reasonText = "Token Invalid"

        errColumns = ['isSuccess','reasonCode','reasonText']
        errData = [(isSuccess,reasonCode,str(e))]

        return toJsonOne(errData,errColumns)

def checkParamDataInput(dataInput,paramList,paramCheckStringList):
    results = []
    for param in paramList:
        if param not in dataInput:
            isSuccess = False
            reasonCode = 500
            reasonText = "Invalid Parameter {}".format(param)

            columns = ['isSuccess','reasonCode','reasonText']
            data = [(isSuccess,reasonCode,reasonText)]

            return toJsonOne(data,columns)

        if (param in paramCheckStringList and isinstance(dataInput[param], str) and len(dataInput[param])==0):
            isSuccess = False
            reasonCode = 500
            reasonText = "Please enter {}".format(param)

            columns = ['isSuccess','reasonCode','reasonText']
            data = [(isSuccess,reasonCode,reasonText)]

            return toJsonOne(data,columns)

class Node(object):
    def __init__(self, data):
        self.data = data
        self.children = []

    def add_child(self, obj):
        self.children.append(obj)

def recur_tree(conn, level, parent, current):
    # print(level, parent.data['table'], current.data['table'])
    if 'sql' not in current.data:
        current.data['sql'] = ''
    if 'hasRow' not in current.data:
        current.data['hasRow'] = False
    if 'paramsValue' not in current.data:
        current.data['paramsValue'] = {}
    if 'params_orig' not in current.data:
        current.data['params_orig'] = current.data['params'].copy()
    else:
        current.data['params'] = current.data['params_orig'].copy()

    for key in parent.data['key'].split(','):
        if key in parent.data['paramsValue']:
            for param in current.data['params']:
                if key == param and current.data['params'][param] == param:
                    current.data['params'][param] = parent.data['paramsValue'][key]

    if len(current.data['sql']) == 0:
        dataCurrent = ()
        columnsCurrent = []
    else:
        cursor = conn.cursor()
        cursor.execute(current.data['sql'],current.data['params'])
        dataCurrent = cursor.fetchall()
        columnsCurrent = [column[0] for column in cursor.description]

    if current.data['hasRow'] and len(dataCurrent) == 0:
        dataCurrent = [['' for x in columnsCurrent]]

    newDataCurrent = []
    listRowCurrent = []
    level = level + 1
    for rowCurrent in dataCurrent:
        paramsValue = {}
        for param in current.data['key'].split(','):
            if param in columnsCurrent:
                paramsValue[param] = rowCurrent[columnsCurrent.index(param)]

        current.data['paramsValue'] = paramsValue
        listRowCurrent = list(rowCurrent)
        if len(current.children) > 0:
            for child in current.children:
                dataChild, columnsChild = recur_tree(conn, level, current, child)
                if 'object' not in child.data or not child.data['object']:
                    childJson = toJson(dataChild,columnsChild)
                else:
                    childJson = toJsonOne(dataChild,columnsChild)
                if child.data['name'] not in columnsCurrent:
                    columnsCurrent.append(child.data['name'])

                listRowCurrent.append(childJson)

                if child.data['name'] + '_delete' not in columnsCurrent:
                    columnsCurrent.append(child.data['name'] + '_delete')

                listRowCurrent.append([])

                if 'selectedVDataTable' in child.data and len(child.data['selectedVDataTable']) > 0:
                    if child.data['selectedVDataTable'] not in columnsCurrent:
                        columnsCurrent.append(child.data['selectedVDataTable'])
                    listRowCurrent.append([])

                if 'searchVDataTable' in child.data and len(child.data['searchVDataTable']) > 0:
                    if child.data['searchVDataTable'] not in columnsCurrent:
                        columnsCurrent.append(child.data['searchVDataTable'])
                    listRowCurrent.append('')

                if 'paginationVDataTable' in child.data and len(child.data['paginationVDataTable']['name']) > 0:
                    if child.data['paginationVDataTable'] not in columnsCurrent:
                        pagination = child.data['paginationVDataTable'].copy()
                        del pagination['name']
                        columnsCurrent.append(child.data['paginationVDataTable']['name'])
                    listRowCurrent.append(pagination)

        newDataCurrent.append(listRowCurrent)

    return newDataCurrent, columnsCurrent

def getGlobalParamMSSQL(globalKey):
    dataInput = request.json

    conn = pymssql.connect(mssql_server,mssql_user,mssql_password,mssql_database)
    cursor = conn.cursor()

    sql = """\
    select parm_value from global_parms
    where global_key = %(global_key)s
    """
    params = {'global_key': globalKey}
    cursor.execute(sql,params)
    data = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    conn.commit()
    cursor.close()

    if len(data) > 0:
        return data[0][0]
    else:
        return ''

def getNextSequenceMSSQL(seqID,tableName,columnName):
    dataInput = request.json

    conn = pymssql.connect(mssql_server,mssql_user,mssql_password,mssql_database)
    cursor = conn.cursor()

    sql = """\
    DECLARE @new_value int;
    EXEC sp_get_next_seq @seq_id = %(seq_id)s,@table_name = %(table_name)s,@column_name = %(column_name)s,@new_value = @new_value OUTPUT;
    select @new_value
    """
    NewValue = 0
    params = {'seq_id': seqID,'table_name': tableName,'column_name': columnName,'new_value': NewValue}
    cursor.execute(sql,params)
    (NewValue,)=cursor.fetchone()

    conn.commit()
    cursor.close()
    return NewValue

def getNextSequenceMySQL(seqID,tableName,columnName):
    dataInput = request.json

    conn = mysql.connect()
    cursor = conn.cursor()

    NewValue = 0

    args = [seqID,tableName,columnName,NewValue]
    cursor.callproc('sp_get_next_seq', args)
    (NewValue,)=cursor.fetchone()

    conn.commit()
    cursor.close()
    return NewValue

def getSequenceNumberMSSQL(seqKey,prefix,year,month,day,formatSeq,formatYear,formatMonth,formatDay,yearOut,monthOut,dayOut,prefixMonth,prefixDay,prefixRunning):
    dataInput = request.json

    conn = pymssql.connect(mssql_server,mssql_user,mssql_password,mssql_database)
    cursor = conn.cursor()

    returnValue = 0

    sql = """\
    DECLARE @ret_value nvarchar(50);
    EXEC sp_get_seq_number @seq_key = %(seq_key)s,@prefix = %(prefix)s,@year = %(year)s,@month = %(month)s,@day = %(day)s,
                            @format_seq = %(format_seq)s,@format_year = %(format_year)s,@format_month = %(format_month)s,@format_day = %(format_day)s,
                            @year_out = %(year_out)s,@month_out = %(month_out)s,@day_out = %(day_out)s,@prefix_month = %(prefix_month)s,
                            @prefix_day = %(prefix_day)s,@prefix_running = %(prefix_running)s,@ret_value = @ret_value OUTPUT
    select @ret_value
    """
    NewValue = 0
    params = {'seq_key': seqKey,'prefix': prefix,'year': year,'month': month,'day': day,
            'format_seq': formatSeq,'format_year': formatYear,'format_month': formatMonth,'format_day': formatDay,
            'year_out': yearOut,'month_out': monthOut,'day_out': dayOut,'prefix_month': prefixMonth,
            'prefix_day': prefixDay,'prefix_running': prefixRunning,'ret_value': returnValue}

    cursor.execute(sql,params)
    (returnValue,)=cursor.fetchone()

    conn.commit()
    cursor.close()
    return returnValue

def getSequenceNumberMySQL(seqKey,prefix,year,month,day,formatSeq,formatYear,formatMonth,formatDay,yearOut,monthOut,dayOut,prefixMonth,prefixDay,prefixRunning):
    dataInput = request.json

    conn = mysql.connect()
    cursor = conn.cursor()

    returnValue = 0

    args = [seqKey,prefix,year,month,day,formatSeq,formatYear,formatMonth,formatDay,yearOut,monthOut,dayOut,prefixMonth,prefixDay,prefixRunning,returnValue]
    cursor.callproc('sp_get_seq_number', args)
    (returnValue,)=cursor.fetchone()

    conn.commit()
    cursor.close()
    return returnValue

@common.route("/readExcel", methods=['POST'])
def readExcel():
    isSuccess = True
    reasonCode = 500
    reasonText = ""
    now = datetime.datetime.now()
    datetimeStr = now.strftime('%Y%m%d_%H%M%S%f')
    filename_tmp = ''
    if request.method == 'POST':
        f = request.files['file']
        filename_tmp = secure_filename('{}_{}'.format(datetimeStr, f.filename))
        f.save(filename_tmp)
    else:
        reasonText = 'Invalid request file'
        errColumns = ['isSuccess','reasonCode','reasonText']
        errData = [(isSuccess,reasonCode,reasonText)]

        return jsonify(toJsonOne(errData,errColumns))

    wb = load_workbook(filename = secure_filename(filename_tmp),data_only=True)
    displayData = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        row_count = sheet.max_row
        column_count = sheet.max_column
        columns = []
        data = []

        for col in range(0, column_count):
            columns.append(col + 1)
            # columns.append(num_to_col_letters(col + 1))
        columns.append('0')
        columns.append('Row_Null')

        is_null = False
        for row in range(0, row_count):
            dataRow = []
            for col in range(0, column_count):
                value = sheet[num_to_col_letters(col + 1)+str(row + 1)].value
                dataRow.append(value)
                if value is 'None':
                    is_null = True

            dataRow.append(row + 1)
            dataRow.append(is_null)
            data.append(tuple(dataRow))

        sheetColumns = ['sheetName','data']
        sheetData = [(sheet_name,toJson(data,columns))]

        displayData.append(toJsonOne(sheetData,sheetColumns))

    os.remove(filename_tmp)
    displayColumns = ['isSuccess','sheet']
    displayData = [(isSuccess,displayData)]
    return jsonify(toJsonOne(displayData,displayColumns))

@common.route("/getMenu", methods=['POST'])
def getMenu():
    now = datetime.datetime.now()
    isSuccess = True
    reasonCode = 200
    reasonText = ""

    dataInput = request.json
    paramList = ['user_id','app_name','sort_type','from_node']
    paramCheckStringList = ['user_id','app_name','sort_type']

    msgError = checkParamDataInput(dataInput,paramList,paramCheckStringList)
    if msgError != None:
        return jsonify(msgError);

    userID = dataInput['user_id']
    appName = dataInput['app_name']
    sortType = dataInput['sort_type']
    fromNode = dataInput['from_node']

    dataInput = request.json

    conn = mysql.connect()
    cursor = conn.cursor()
    args = [userID,appName,sortType,fromNode]
    cursor.callproc('sp_pruned_menu_wrapper', args)

    data = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    cursor.close()

    displayColumns = ['isSuccess','data']
    displayData = [(isSuccess,toJson(data,columns))]

    return jsonify(toJsonOne(displayData,displayColumns))
