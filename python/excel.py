# -*- coding: utf-8 -*-

from flask import Flask, session, redirect, url_for
from flask import Flask, request, jsonify, current_app, abort, send_from_directory, send_file
from flask_cors import CORS, cross_origin
from flask import Blueprint
from flaskext.mysql import MySQL
from pymongo import MongoClient
import json
from bson import BSON
from bson import json_util
import os
import pymssql
import datetime
import re
import random
import math
import base64
import codecs
from flask import Flask, request, make_response
from pyreportjasper import JasperPy
from openpyxl import load_workbook
from openpyxl import Workbook
from werkzeug import secure_filename

from db_config import *
from common import *
excel = Blueprint('excel', __name__)

@excel.route("/excel/testExcel", methods=['POST'])
def testExcel():
    now = datetime.datetime.now()
    isSuccess = True
    reasonCode = 200
    reasonText = ""

    dataInput = request.json
    paramList = ['data_json']
    paramCheckStringList = ['data_json']

    msgError = checkParamDataInput(dataInput,paramList,paramCheckStringList)
    if msgError != None:
        return jsonify(msgError);

    dataJson = dataInput['data_json']

    now = datetime.datetime.now()
    datetimeStr = now.strftime('%Y%m%d_%H%M%S%f')
    filename_tmp = secure_filename('{}_{}'.format(datetimeStr, 'Template_Poll.xlsx'))
    wb = load_workbook('Template\Template_Poll.xlsx')

    if len(dataJson) > 0:
        sheet = wb['แบบฟอร์มสำหรับจัดเก็บข้อมูล อปท']
        sheet['D2'] = dataJson[0]['json_taxYearBE']
        sheet['D3'] = dataJson[0]['json_changwatsInput']
        sheet['D4'] = dataJson[0]['json_subDistrictHeaderInput']
        sheet['D5'] = dataJson[0]['json_DistrictHeaderInput']

        offset = 9
        i = 0
        for row in dataJson:
            sheet['A'+str(offset + i)] = row['json_No']
            sheet['B'+str(offset + i)] = row['json_CodeOwner']
            sheet['C'+str(offset + i)] = row['json_DistrictHeaderInput']
            sheet['D'+str(offset + i)] = row['json_TTaxpayerStatus']
            sheet['E'+str(offset + i)] = row['json_TTypeUseLand']
            sheet['F'+str(offset + i)] = row['json_TypeOwner']
            sheet['G'+str(offset + i)] = row['json_TypeUseLands']
            sheet['H'+str(offset + i)] = row['json_ParcelNumber']
            sheet['I'+str(offset + i)] = row['json_DeedNumber']
            sheet['J'+str(offset + i)] = row['json_SheetNumber']
            sheet['K'+str(offset + i)] = row['json_SpaceAreaSqw']
            sheet['L'+str(offset + i)] = row['json_CostEstimateSpaceAreaSQ']
            sheet['M'+str(offset + i)] = row['json_CostEstimate']
            sheet['N'+str(offset + i)] = row['json_PlantType']
            sheet['O'+str(offset + i)] = row['json_plantFloor']
            sheet['P'+str(offset + i)] = row['json_plantYear']
            sheet['Q'+str(offset + i)] = row['json_SpacePlantingAreaSqm']
            sheet['R'+str(offset + i)] = row['json_CostEstimateSpacePlantingAreaSqm']
            sheet['S'+str(offset + i)] = row['json_CostEstimateSpacePlantingArea']
            sheet['T'+str(offset + i)] = row['json_DepreciationRate']
            sheet['U'+str(offset + i)] = row['json_Depreciation']
            sheet['V'+str(offset + i)] = row['json_CostEstimateSpacePlantingAreaDepreciation']
            sheet['W'+str(offset + i)] = row['json_TypeUseMix']
            sheet['X'+str(offset + i)] = row['json_SpacePlantingAreaSqmMix']
            sheet['Y'+str(offset + i)] = row['json_CostEstimateMix']
            sheet['Z'+str(offset + i)] = row['json_CostEstimateSpacePlantingAreaMix']
            sheet['AA'+str(offset + i)] = row['json_CostEstimateLandAndBuilding']
            sheet['AB'+str(offset + i)] = row['json_TaxRate']
            sheet['AC'+str(offset + i)] = row['json_AmountTotal']
            sheet['AD'+str(offset + i)] = row['json_Allowance']
            sheet['AE'+str(offset + i)] = row['json_TaxOlderLandAndProperty']
            sheet['AF'+str(offset + i)] = row['json_TaxOlderLocal']
            sheet['AG'+str(offset + i)] = row['json_AmountTotalNet']
            sheet['AH'+str(offset + i)] = row['json_TaxActual']
            sheet['AI'+str(offset + i)] = row['json_remark']
            # j = 1
            # for col in row:
            #     sheet[num_to_col_letters(j)+str(offset + i)] = row[col]
            #     j = j + 1
            i = i + 1

    wb.save(filename_tmp)

    with open(os.path.dirname(os.path.abspath(__file__)) + '/' + filename_tmp, "rb") as f:
        encoded_string = base64.b64encode(f.read())

    os.remove(filename_tmp)

    displayColumns = ['isSuccess','reasonCode','reasonText','excel_base64']
    displayData = [(isSuccess,reasonCode,reasonText,encoded_string)]

    return jsonify(toJsonOne(displayData,displayColumns))

@excel.route("/excel/testCustoms", methods=['POST'])
def testCustoms():
    try:
        isSuccess = True
        reasonCode = 500
        reasonText = ""
        now = datetime.datetime.now()
        datetimeStr = now.strftime('%Y%m%d_%H%M%S%f')
        filename_tmp = ''
        if request.method == 'POST':
            f = request.files['file']
            filename_tmp = secure_filename('{}_{}'.format(datetimeStr, f.filename))
            f.save(filename_tmp)
        else:
            reasonText = 'Invalid request file'
            errColumns = ['isSuccess','reasonCode','reasonText']
            errData = [(isSuccess,reasonCode,reasonText)]

            return jsonify(toJsonOne(errData,errColumns))

        wb = load_workbook(filename = secure_filename(filename_tmp),data_only=True)
        displayData = []

        errExcelList = []
        transData = []
        conn = mysql.connect()

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            row_count = sheet.max_row
            column_count = sheet.max_column
            columns = []
            data = []
            cursor = conn.cursor()
            sql = """\
            select form_survey_master.Form_Survey_id,form_survey_master.Form_Survey_desc,
            form_survey_attribute_master.Attribute_Master_Attribute_id,form_survey_attribute_master.Cell_Excel,
            attribute_master.Attribute_Name,attribute_master.Data_type,attribute_master.Value_List,
            attribute_master.minimum_value,attribute_master.maximum_value,attribute_master.expression
            from Form_Survey_Master inner join Form_Survey_Attribute_Master on
            Form_Survey_Master.Form_Survey_id = Form_Survey_Attribute_Master.Form_Survey_Master_Form_Survey_id inner join Attribute_Master on
            Form_Survey_Attribute_Master.Attribute_Master_Attribute_id = Attribute_Master.Attribute_id
            where form_survey_master.Form_Survey_desc = %(Form_Survey_desc)s
            """
            params = {'Form_Survey_desc': sheet["A1"].value}
            cursor.execute(sql,params)
            formSurveyData=cursor.fetchall()
            formSurveycolumns = [column[0] for column in cursor.description]

            for formSurvey in toJson(formSurveyData,formSurveycolumns):
                valueExcel = sheet[formSurvey['Cell_Excel']].value
                if formSurvey['Data_type'] == 'String':
                    if not formSurvey['Value_List'] is None and len(formSurvey['Value_List']) > 0:
                        if not existsListString(formSurvey['Value_List'].split("|"),valueExcel):
                            err = {
                                "sheetName":sheet_name,
                                "cell":formSurvey['Cell_Excel'],
                                "error":'{} not in {}'.format(valueExcel,formSurvey['Value_List'])
                                }
                            errExcelList.append(err)
                elif formSurvey['Data_type'] == 'Integer':
                    if  not type(valueExcel) is int:
                        err = {
                            "sheetName":sheet_name,
                            "cell":formSurvey['Cell_Excel'],
                            "error":'{} type isnot Integer'.format(valueExcel)
                            }
                        errExcelList.append(err)
                    elif not formSurvey['minimum_value'] is None and len(formSurvey['minimum_value']) > 0:
                        try:
                           minValue = int(formSurvey['minimum_value'])
                           if  valueExcel < minValue:
                               err = {
                                   "sheetName":sheet_name,
                                   "cell":formSurvey['Cell_Excel'],
                                   "error":'{} must greater than {}'.format(valueExcel,minValue)
                                   }
                               errExcelList.append(err)
                        except ValueError:
                            pass
                        if not formSurvey['maximum_value'] is None and len(formSurvey['maximum_value']) > 0:
                            try:
                               maxValue = int(formSurvey['maximum_value'])
                               if  valueExcel > maxValue:
                                   err = {
                                       "sheetName":sheet_name,
                                       "cell":formSurvey['Cell_Excel'],
                                       "error":'{} must less than {}'.format(valueExcel,maxValue)
                                       }
                                   errExcelList.append(err)
                            except ValueError:
                                pass

                if len(formSurvey['expression']) > 0:
                    expression = formSurvey['expression'].replace('@value',str(valueExcel))
                    valueExcel = eval(expression)
                    # sheet['A7'] = expression
                    # valueExcel = sheet['A7'].value

                    # book = Workbook()
                    # sheet = book.active
                    #
                    # sheet['A1'] = expression
                    # book.save("sample.xlsx")

                    # book = xlwt.Workbook()
                    # ws = book.add_sheet('Sheet1')
                    #
                    # ws.write(2, 2, xlwt.Formula(expression))
                    #
                    # book.save('sample.xls')


                    # book = load_workbook('sample.xlsx',data_only=True)
                    # sheet = book.active
                    # valueExcel = sheet['B2'].value

                trans = {
                    "Form_Survey_id":formSurvey['Form_Survey_id'],
                    "Attribute_id":formSurvey['Attribute_Master_Attribute_id'],
                    "value":valueExcel
                }

                transData.append(trans)

        if len(errExcelList) > 0:
            isSuccess = False
            reasonCode = 500
            errColumns = ['isSuccess','reasonCode','reasonText','errExcel']
            errData = [(isSuccess,reasonCode,'Found Error',errExcelList)]

            return jsonify(toJsonOne(errData,errColumns))
        else:
            for trans in transData:
                transSkey = getNextSequenceMySQL('form_survey_transaction','form_survey_transaction','Transaction_id')
                # transID = getSequenceNumberMySQL('form_survey_transaction','TRN',now.year,now.month,now.day,'00000000','00','00','00','Y','Y','Y','','','-')
                sql = """\
                insert into form_survey_transaction
                (Transaction_id,Form_Survey_Master_Form_Survey_id,Attribute_Master_Attribute_id,value)
                values(%(Transaction_id)s,%(Form_Survey_Master_Form_Survey_id)s,%(Attribute_Master_Attribute_id)s,%(value)s)
                """
                params = {'Transaction_id': transSkey,'Form_Survey_Master_Form_Survey_id': trans['Form_Survey_id'],'Attribute_Master_Attribute_id': trans['Attribute_id'],'value': trans['value']}

                cursor.execute(sql,params)

            conn.commit()
            displayColumns = ['isSuccess','reasonCode','reasonText']
            displayData = [(isSuccess,reasonCode,'Success')]

            return jsonify(toJsonOne(displayData,displayColumns))
    except Exception as e:
        conn.rollback()
        isSuccess = False
        reasonCode = 500

        errColumns = ['isSuccess','reasonCode','reasonText']
        errData = [(isSuccess,reasonCode,str(e))]

        return jsonify(toJsonOne(errData,errColumns))
    finally:
        # Check File Exists
        os.remove(filename_tmp)

@excel.route("/excel/retrieveCustoms", methods=['POST'])
def retrieveCustoms():
    now = datetime.datetime.now()
    isSuccess = True
    reasonCode = 200
    reasonText = ""

    dataInput = request.json
    paramList = ['province']
    paramCheckStringList = []

    msgError = checkParamDataInput(dataInput,paramList,paramCheckStringList)
    if msgError != None:
        return jsonify(msgError);

    province = dataInput['province']

    conn = mysql.connect()
    cursor = conn.cursor()
    sql = """\
    select province,province_name,sum(import_duty) as sum_import_duty,sum(export_duty) as sum_export_duty,
    sum(fee) as sum_fee,sum(income) as sum_income
    from province_customs
    where province like %(province)s
    group by province
    """
    params = {'province': province}
    cursor.execute(sql,params)

    data = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    conn.commit()
    cursor.close()
    displayColumns = ['isSuccess','data']
    displayData = [(isSuccess,toJson(data,columns))]

    return jsonify(toJsonOne(displayData,displayColumns))
